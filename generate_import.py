"""
Generate SQL import for staging DB:
- Tower table (SUTT 150kV + SUTET 500kV from KOORDINAT TOWER sheet)
- JalurKML table (SUTT/SUTET spans + SKTT from KOORDINAT SKTT sheet)
- Laporan table (PPL kerawanan from PPL file)
"""
import openpyxl
import re
import json
import sys
from datetime import datetime

EXCEL_TOWER = '/Volumes/project-danu/bornworks/frontend-sitower/Data Koordinat Jaringan.xlsx'
EXCEL_PPL   = '/Volumes/project-danu/bornworks/frontend-sitower/Data PPL dan Kerawanan ULTG Durikosambi-3.xlsx'

lines = []
def sql(s): lines.append(s)

def esc(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"

def ts():
    return "NOW()"

# ── 1. TOWER table ────────────────────────────────────────────────────────────
print("[1] Reading KOORDINAT TOWER...", file=sys.stderr)
wb_t = openpyxl.load_workbook(EXCEL_TOWER, read_only=True)
ws_tower = wb_t['KOORDINAT TOWER']

# Route-code → tipe mapping (for JalurKML)
span_routes = {}  # {(route_code, tipe): [(lat, lng), ...]}

tower_rows = []
for row in ws_tower.iter_rows(min_row=2, values_only=True):
    tipe_raw = str(row[3]).strip() if row[3] else ''
    name     = str(row[2]).strip() if row[2] else ''
    lokasi   = str(row[1]).strip() if row[1] else ''
    lat_raw  = row[4]
    lng_raw  = row[5]

    try:
        lat = float(str(lat_raw).replace(',', '.')) if lat_raw else None
        lng = float(str(lng_raw).replace(',', '.')) if lng_raw else None
    except:
        lat = lng = None

    if not name or lat is None or lng is None:
        continue

    # --- Tower inserts ---
    if tipe_raw in ('TOWER 150', 'TOWER 500'):
        tegangan = '150kV' if tipe_raw == 'TOWER 150' else '500kV'
        tipe_db  = 'SUTT'  if tipe_raw == 'TOWER 150' else 'SUTET'
        # Extract nomor urut from name e.g. #0001
        m = re.search(r'#(\w+)$', name)
        nomor_str = m.group(1) if m else None
        try:
            nomor_urut = int(nomor_str) if nomor_str and nomor_str.isdigit() else None
        except:
            nomor_urut = None
        tower_rows.append((name, lat, lng, tegangan, tipe_db, lokasi, nomor_urut))

    # --- Span rows for JalurKML paths ---
    elif tipe_raw in ('SPAN SUTT', 'SPAN SUTET'):
        tipe_db = 'SUTT' if tipe_raw == 'SPAN SUTT' else 'SUTET'
        m = re.search(r'(?:SUTT|SUTET)\s+\d+kV\s+([A-Z0-9+\-]+)\s+#(\w+)', name)
        if m:
            code = m.group(1)
            num_str = m.group(2)
            # Convert span label to sort key: "0001"→1, "021A"→210, "0034-GT"→340
            try:
                sort_key = int(re.sub(r'[^0-9]', '', num_str) or '9999')
            except:
                sort_key = 9999
            key = (code, tipe_db)
            if key not in span_routes:
                span_routes[key] = []
            span_routes[key].append((sort_key, lat, lng))

sql("-- ═══ TOWERS ═══")
sql("BEGIN;")
sql('TRUNCATE "Tower" CASCADE;')

for (name, lat, lng, tegangan, tipe, lokasi, nomor) in tower_rows:
    nomor_sql = str(nomor) if nomor else 'NULL'
    sql(f'''INSERT INTO "Tower" (id, nama, lat, lng, tegangan, tipe, kondisi, lokasi, "nomorUrut", "statusKerawanan", radius, "createdAt", "updatedAt")
VALUES ({esc(name)}, {esc(name)}, {lat}, {lng}, {esc(tegangan)}, {esc(tipe)}, 'normal', {esc(lokasi) if lokasi else 'NULL'}, {nomor_sql}, 'aman', 100, NOW(), NOW())
ON CONFLICT (id) DO NOTHING;''')

sql("COMMIT;")
print(f"[1] {len(tower_rows)} towers queued", file=sys.stderr)

# ── 2. JalurKML — SUTT/SUTET from SPAN rows ──────────────────────────────────
print("[2] Building SUTT/SUTET JalurKML paths...", file=sys.stderr)
sql("\n-- ═══ JalurKML SUTT/SUTET ═══")
sql("BEGIN;")

SUTT_COLOR  = '#0000FF'
SUTET_COLOR = '#FF0000'

for (code, tipe), pts in sorted(span_routes.items()):
    # Sort by span number, then deduplicate
    pts.sort(key=lambda x: x[0])
    seen = set()
    unique_pts = []
    for (sort_key, lat, lng) in pts:
        k = (round(lat,6), round(lng,6))
        if k not in seen:
            seen.add(k)
            unique_pts.append({'lat': lat, 'lng': lng})

    warna = SUTT_COLOR if tipe == 'SUTT' else SUTET_COLOR
    nama  = f"{tipe} {code}"
    path_json = json.dumps(unique_pts)
    sql(f'''INSERT INTO "JalurKML" (nama, tipe, warna, path, "createdAt", "updatedAt")
VALUES ({esc(nama)}, {esc(tipe)}, {esc(warna)}, {esc(path_json)}::jsonb, NOW(), NOW())
ON CONFLICT DO NOTHING;''')

sql("COMMIT;")
print(f"[2] {len(span_routes)} SUTT/SUTET jalur queued", file=sys.stderr)

# ── 3. JalurKML — SKTT from KOORDINAT SKTT sheet ─────────────────────────────
print("[3] Reading KOORDINAT SKTT...", file=sys.stderr)
ws_sktt = wb_t['KOORDINAT SKTT']

sktt_routes = {}  # {ruas: {sort_key: {lat, lng}}}
for row in ws_sktt.iter_rows(min_row=2, values_only=True):
    ruas    = str(row[2]).strip() if row[2] else ''
    bay     = str(row[3]).strip() if row[3] else ''  # BAY_PENGHANTAR e.g. "JOINT SKTT 150kV KBANG-DKSBI 1 #0003"
    lat_raw = row[14]
    lng_raw = row[15]

    if not ruas or ruas == 'RUAS':
        continue

    try:
        lat = float(str(lat_raw).replace(',', '.')) if lat_raw else None
        lng = float(str(lng_raw).replace(',', '.')) if lng_raw else None
    except:
        lat = lng = None

    if lat is None or lng is None or lat == 0 or lng == 0:
        continue

    # Extract joint number for sorting, e.g. "#0003" → 3
    m = re.search(r'#(\w+)', bay)
    try:
        sort_key = int(re.sub(r'[^0-9]', '', m.group(1)) or '9999') if m else 9999
    except:
        sort_key = 9999

    coord_key = f"{round(lat,7)},{round(lng,7)}"
    if ruas not in sktt_routes:
        sktt_routes[ruas] = {}
    # Keep first occurrence of each unique coord, with sort_key for ordering
    if coord_key not in sktt_routes[ruas]:
        sktt_routes[ruas][coord_key] = (sort_key, lat, lng)

sql("\n-- ═══ JalurKML SKTT ═══")
sql("BEGIN;")

SKTT_COLOR = '#800080'

for ruas, pts_map in sorted(sktt_routes.items()):
    # Sort by joint number
    pts = sorted(pts_map.values(), key=lambda x: x[0])
    pts = [{'lat': p[1], 'lng': p[2]} for p in pts]
    if not pts:
        continue
    path_json = json.dumps(pts)
    sql(f'''INSERT INTO "JalurKML" (nama, tipe, warna, path, "createdAt", "updatedAt")
VALUES ({esc(ruas)}, 'SKTT', {esc(SKTT_COLOR)}, {esc(path_json)}::jsonb, NOW(), NOW())
ON CONFLICT DO NOTHING;''')

sql("COMMIT;")
print(f"[3] {len(sktt_routes)} SKTT jalur queued", file=sys.stderr)

# ── 4. Laporan (PPL kerawanan) ─────────────────────────────────────────────────
print("[4] Reading PPL kerawanan...", file=sys.stderr)
wb_ppl = openpyxl.load_workbook(EXCEL_PPL, read_only=True)
ws_ppl = wb_ppl['Sheet1']

# Build tower lookup: (ruas_code, tower_num) → tower_id
# Tower name format: "TOWER SUTT 150kV CKRBR-TGBRU+CKRBR-ITS #0026"
# PPL RUAS: "SUTT 150kV CKRBR-TGBRU+CKRBR-ITS"
def normalize_code(code):
    return (code.upper()
            .replace('DKSB7', 'DKSBI')
            .replace('KMBGN', 'KMBNG')
            .replace('7', 'I'))

tower_lookup = {}       # (norm_code, tower_num) → tower_name
tower_lookup_raw = {}   # (raw_code, tower_num) → tower_name
tower_by_num = {}       # tower_num → [(norm_code, tower_name)]

for (name, lat, lng, tegangan, tipe, lokasi, nomor) in tower_rows:
    # Format 1: "TOWER SUTT 150kV DUKSB-CGKRG #0001"  (voltage before code)
    # Format 2: "TOWER SUTET KMBGN-DKSBI 500kV #PA01" (voltage after code)
    m = re.search(r'(?:SUTT|SUTET)\s+(?:\d+kV\s+)?([A-Z0-9+\-]+?)(?:\s+\d+kV)?\s+#(\w+)', name)
    if m:
        raw_code = m.group(1).strip()
        num = '#' + m.group(2)
        norm = normalize_code(raw_code)
        tower_lookup[(norm, num)] = name
        tower_lookup_raw[(raw_code, num)] = name
        if num not in tower_by_num:
            tower_by_num[num] = []
        tower_by_num[num].append((norm, name))

def match_tower(ruas_raw, tower_num):
    """Try to find tower ID from PPL RUAS + NO. TOWER"""
    if not ruas_raw or not tower_num:
        return None
    ruas_raw = str(ruas_raw).strip()
    tower_num = str(tower_num).strip()

    # Skip TRS/SKTT entries — no Tower record for these
    if str(ruas_raw).upper().startswith('TRS'):
        return None

    m = re.search(r'(?:SUTT|SUTET)\s+(?:\d+kV\s+)?([A-Z0-9+\-]+)', ruas_raw, re.I)
    if not m:
        return None
    raw_code = m.group(1).strip().upper()
    norm_code = normalize_code(raw_code)

    # Direct normalized match
    key = (norm_code, tower_num)
    if key in tower_lookup:
        return tower_lookup[key]

    # Raw match
    if (raw_code, tower_num) in tower_lookup_raw:
        return tower_lookup_raw[(raw_code, tower_num)]

    # Fuzzy: same tower_num, code overlap
    candidates = tower_by_num.get(tower_num, [])
    for (tc, tid) in candidates:
        # Check if enough of the route prefix matches
        parts_query = set(norm_code.replace('+', '-').split('-'))
        parts_tower = set(tc.replace('+', '-').split('-'))
        overlap = parts_query & parts_tower
        if len(overlap) >= 2:  # at least 2 matching segments
            return tid

    return None

def map_jenis(klasifikasi):
    k = str(klasifikasi or '').lower().strip()
    if 'pihak lain' in k or k == 'ppl': return 'pekerjaan_pihak_lain'
    if 'layangan' in k: return 'layangan'
    if 'kebakaran' in k: return 'kebakaran'
    if 'pencurian' in k: return 'pencurian'
    if 'pemanfaatan' in k: return 'pemanfaatan_lahan'
    return 'pekerjaan_pihak_lain'

def map_status(status_raw):
    s = str(status_raw or '').lower().strip()
    if 'selesai' in s: return 'selesai'
    if 'berlangsung' in s: return 'berlangsung'
    if 'aktifitas' in s or 'tidak ada' in s: return 'selesai'
    return 'berlangsung'

def map_level(status_raw):
    s = str(status_raw or '').lower().strip()
    if 'kritis' in s: return 'kritis'
    if 'sedang' in s: return 'sedang'
    return 'rendah'

# Default pelapor
DEFAULT_PELAPOR_ID = 'pegawai-teknisi-001'

# Pegawai name → id cache (from known data)
PEGAWAI_MAP = {
    'apippudin': 'cmox3g6n10000190qw7ktvua4',
    'fauzi': 'cmox3g6nd0003190qk4qgozak',
    'rudi': 'cmox3g6nj0009190qmsw6kzno',
    'andika': 'cmox3g6nl000b190qtr8qby0y',
    'idham': 'cmox3g6nt000h190qhb13e836',
    'ahmad fadli': 'cmox3g6o3000o190qgwwux37v',
    'febrian': 'cmox3g6o5000q190q3q9x1l0s',
    'ahmad fauzi': 'cmowpszl20002yz0qu60rmujl',
}

sql("\n-- ═══ Laporan PPL ═══")
sql("BEGIN;")

laporan_ok = 0
laporan_skip = 0

import uuid

for row in ws_ppl.iter_rows(min_row=4, values_only=True):
    no_raw = row[0]
    if not no_raw:
        continue
    try:
        float(str(no_raw))
    except:
        continue

    ruas_raw     = row[1]
    tower_no_raw = row[2]  # NO. TOWER e.g. "#0026"
    span_raw     = row[3]  # SPAN e.g. "T25 - T26"
    uraian       = row[4]  # URAIAN PEKERJAAN
    klasifikasi  = row[5]  # KLASIFIKASI
    progres      = row[6]  # PROGRES
    status_raw   = row[7]  # STATUS
    pengendalian = row[8]  # PENGENDALIAN
    pihak_lain   = row[9]  # PIHAK LAIN
    contact      = row[10] # CONTACT PERSON
    petugas      = row[11] # PETUGAS LW

    tower_id = match_tower(ruas_raw, str(tower_no_raw).strip() if tower_no_raw else None)
    if not tower_id:
        laporan_skip += 1
        continue

    # Pelapor
    petugas_clean = str(petugas or '').strip().lower().lstrip()
    pelapor_id = PEGAWAI_MAP.get(petugas_clean, DEFAULT_PELAPOR_ID)

    jenis    = map_jenis(klasifikasi)
    status   = map_status(status_raw)
    level    = map_level(status_raw)
    deskripsi = str(uraian or '').strip()
    ket_parts = []
    if pengendalian: ket_parts.append(str(pengendalian).strip())
    if pihak_lain:   ket_parts.append(f"Pihak Lain: {str(pihak_lain).strip()}")
    if contact:      ket_parts.append(f"Contact: {str(contact).strip()}")
    if span_raw:     ket_parts.append(f"Span: {str(span_raw).strip()}")
    keterangan = '\n'.join(ket_parts) if ket_parts else None

    lid = str(uuid.uuid4()).replace('-', '')[:25]

    sql(f'''INSERT INTO "Laporan" (id, "towerId", "pelaporId", deskripsi, status, tanggal, keterangan, "jenisGangguan", "levelRisiko", "createdAt", "updatedAt")
VALUES ({esc(lid)}, {esc(tower_id)}, {esc(pelapor_id)}, {esc(deskripsi)}, {esc(status)}, NOW(), {esc(keterangan)}, {esc(jenis)}, {esc(level)}, NOW(), NOW());''')
    laporan_ok += 1

sql("COMMIT;")
print(f"[4] {laporan_ok} laporan queued, {laporan_skip} skipped (tower not found)", file=sys.stderr)

# ── Output ────────────────────────────────────────────────────────────────────
print('\n'.join(lines))
